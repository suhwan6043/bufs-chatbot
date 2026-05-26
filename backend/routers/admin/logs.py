"""대화 로그 조회/내보내기 API — pages/admin.py logs 섹션 이식."""

import io
import json
from collections import Counter
from datetime import date, timedelta

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse

from backend.routers.admin.auth import require_admin

router = APIRouter()

_INTENT_LABELS = {
    "GRADUATION_REQ": "졸업요건", "REGISTRATION": "수강신청",
    "SCHEDULE": "학사일정", "COURSE_INFO": "교과목", "MAJOR_CHANGE": "전과",
    "ALTERNATIVE": "대안/선택", "GENERAL": "일반",
    "LEAVE_OF_ABSENCE": "학적변동", "EARLY_GRADUATION": "조기졸업",
    "SCHOLARSHIP": "장학금", "CONTACT": "연락처",
}


@router.get("/logs/dates")
async def get_log_dates(_=Depends(require_admin)):
    """사용 가능한 로그 날짜 목록."""
    from app.logging import ChatLogger
    logger = ChatLogger()
    dates = logger.list_dates()
    return {"dates": [d.isoformat() for d in dates]}


@router.get("/logs")
async def get_logs(
    log_date: str = Query(None, description="날짜 (YYYY-MM-DD). 없으면 전체"),
    intent: str = Query(None, description="인텐트 필터"),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    _=Depends(require_admin),
):
    """대화 로그 조회."""
    from app.logging import ChatLogger

    logger = ChatLogger()

    if log_date:
        try:
            parts = log_date.split("-")
            d = date(int(parts[0]), int(parts[1]), int(parts[2]))
            entries = logger.read(d)
        except Exception:
            entries = []
    else:
        entries = logger.read_all()

    # 인텐트 필터
    if intent:
        entries = [e for e in entries if e.get("intent") == intent]

    total = len(entries)

    # KPI
    today_count = len(logger.read(date.today()))
    avg_ms = (sum(e.get("duration_ms", 0) for e in entries) / total) if total else 0
    intent_counter = Counter(e.get("intent", "") for e in entries if e.get("intent"))
    top_intent_raw = intent_counter.most_common(1)
    top_intent = _INTENT_LABELS.get(top_intent_raw[0][0], top_intent_raw[0][0]) if top_intent_raw else "-"

    # 페이지네이션 + 최신순 정렬
    sorted_entries = sorted(entries, key=lambda x: x.get("timestamp", ""), reverse=True)
    paged = sorted_entries[offset:offset + limit]

    return {
        "total": total,
        "today_count": today_count,
        "avg_duration_ms": round(avg_ms),
        "top_intent": top_intent,
        "entries": paged,
    }


@router.get("/logs/export/csv")
async def export_csv(
    log_date: str = Query(None),
    _=Depends(require_admin),
):
    """CSV 내보내기."""
    from app.logging import ChatLogger

    logger = ChatLogger()
    if log_date:
        try:
            parts = log_date.split("-")
            entries = logger.read(date(int(parts[0]), int(parts[1]), int(parts[2])))
        except Exception:
            entries = logger.read_all()
    else:
        entries = logger.read_all()

    import csv
    buf = io.StringIO()
    fields = ["timestamp", "session_id", "student_id", "intent", "question", "answer", "duration_ms", "rating"]
    writer = csv.DictWriter(buf, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    for e in entries:
        writer.writerow(e)

    output = buf.getvalue().encode("utf-8-sig")
    filename = f"camchat_logs_{date.today().isoformat()}.csv"

    return StreamingResponse(
        io.BytesIO(output),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/logs/export/jsonl")
async def export_jsonl(
    log_date: str = Query(None),
    _=Depends(require_admin),
):
    """JSONL 내보내기."""
    from app.logging import ChatLogger

    logger = ChatLogger()
    if log_date:
        try:
            parts = log_date.split("-")
            entries = logger.read(date(int(parts[0]), int(parts[1]), int(parts[2])))
        except Exception:
            entries = logger.read_all()
    else:
        entries = logger.read_all()

    output = "\n".join(json.dumps(e, ensure_ascii=False) for e in entries).encode("utf-8")
    filename = f"camchat_logs_{date.today().isoformat()}.jsonl"

    return StreamingResponse(
        io.BytesIO(output),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/logs/export/xlsx")
async def export_xlsx(
    log_date: str = Query(None),
    _=Depends(require_admin),
):
    """엑셀(.xlsx) 내보내기 — 헤더 스타일·자동 필터·첫 행 고정.

    학사지원팀이 수합·분석하기 좋도록 한글 헤더와 칼럼 너비를 사전 설정.
    권한: JWT 로그인만 통과하면 OK (`require_admin`).
    """
    from app.logging import ChatLogger
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    logger = ChatLogger()
    if log_date:
        try:
            parts = log_date.split("-")
            entries = logger.read(date(int(parts[0]), int(parts[1]), int(parts[2])))
        except Exception:
            entries = logger.read_all()
    else:
        entries = logger.read_all()

    # 최신순 정렬 (대시보드/로그 페이지와 동일 순서)
    entries = sorted(entries, key=lambda x: x.get("timestamp", ""), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "대화 로그"

    # 헤더
    headers = ["시간", "세션ID", "학번", "Intent", "질문", "답변", "응답(ms)", "만족도"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 데이터 행 — Intent 한글 라벨 적용
    body_align = Alignment(vertical="top", wrap_text=True)
    for e in entries:
        intent = _INTENT_LABELS.get(e.get("intent", ""), e.get("intent", ""))
        ws.append([
            e.get("timestamp", ""),
            e.get("session_id", ""),
            e.get("student_id", "") or "",
            intent,
            e.get("question", ""),
            e.get("answer", ""),
            e.get("duration_ms", 0),
            e.get("rating", "") if e.get("rating") is not None else "",
        ])

    # 본문 줄바꿈 (질문·답변 칼럼이 길어도 줄바꿈으로 표시)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = body_align

    # 칼럼 너비
    widths = [20, 25, 12, 14, 50, 70, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 자동 필터 + 첫 행 고정
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    suffix = log_date if log_date else date.today().isoformat()
    filename = f"camchat_logs_{suffix}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
