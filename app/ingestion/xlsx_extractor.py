"""
XLSX 추출기 - openpyxl을 이용해 Excel 파일에서 시트별 데이터를 추출합니다.

반환값: List[PageContent]
  - 시트 1개 = PageContent 1개 (page_number = 시트 인덱스 + 1)
  - 데이터 행은 마크다운 테이블로 변환하여 tables 필드에 저장
  - 시트 이름은 text 필드 앞에 헤더로 삽입
"""

import logging
from pathlib import Path
from typing import List, Optional

from app.models import PageContent

logger = logging.getLogger(__name__)

# 최소 유효 행 수 (이 미만이면 시트 스킵)
_MIN_ROWS = 1
# 마크다운 테이블 셀 최대 길이 (길면 잘라서 표시)
_MAX_CELL_LEN = 200


def _cell_str(value) -> str:
    """셀 값을 문자열로 변환합니다. None은 빈 문자열."""
    if value is None:
        return ""
    return str(value).strip().replace("\n", " ")[:_MAX_CELL_LEN]


def _rows_to_markdown(rows: list[list[str]]) -> str:
    """
    2D 문자열 배열을 마크다운 테이블로 변환합니다.
    첫 번째 행을 헤더로 취급합니다.
    """
    if not rows:
        return ""

    # 열 수 통일 (최대 열 기준 패딩)
    max_cols = max(len(r) for r in rows)
    padded = [r + [""] * (max_cols - len(r)) for r in rows]

    lines = []
    for i, row in enumerate(padded):
        lines.append("| " + " | ".join(row) + " |")
        if i == 0:
            lines.append("|" + "|".join(["---"] * max_cols) + "|")

    return "\n".join(lines)


class XlsxExtractor:
    """
    XLSX/XLS 파일에서 시트별 데이터를 추출하는 클래스.

    사용법:
        extractor = XlsxExtractor()
        pages = extractor.extract("path/to/file.xlsx")
    """

    def extract(self, path: str) -> List[PageContent]:
        """
        XLSX 파일을 읽어 시트별 PageContent 목록을 반환합니다.

        Args:
            path: XLSX 파일 경로

        Returns:
            List[PageContent] (빈 파일이면 [])
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl이 설치되지 않았습니다: pip install openpyxl")
            return []

        try:
            # data_only=True: 수식 대신 계산된 값 사용
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            logger.error("XLSX 열기 실패 [%s]: %s", path, e)
            return []

        filename = Path(path).name
        pages: List[PageContent] = []

        for sheet_idx, ws in enumerate(wb.worksheets):
            sheet_name = ws.title or f"Sheet{sheet_idx + 1}"

            # 유효 행 수집 (빈 행 제거)
            data_rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                str_row = [_cell_str(c) for c in row]
                # 완전히 빈 행 스킵
                if any(c for c in str_row):
                    data_rows.append(str_row)

            if len(data_rows) < _MIN_ROWS:
                logger.debug("빈 시트 건너뜀: %s / %s", filename, sheet_name)
                continue

            table_md = _rows_to_markdown(data_rows)

            # text 필드: 시트 이름 컨텍스트
            text_header = f"[파일: {filename}] [시트: {sheet_name}] (행 {len(data_rows)}개)"

            pages.append(PageContent(
                page_number=sheet_idx + 1,
                text=text_header,
                tables=[table_md] if table_md else [],
                source_file=str(path),
            ))

            logger.debug(
                "시트 추출: %s / %s → %d행",
                filename, sheet_name, len(data_rows),
            )

        wb.close()

        if not pages:
            logger.warning("XLSX에서 추출된 내용 없음: %s", filename)

        logger.info("XLSX 추출 완료: %s → %d시트", filename, len(pages))
        return pages
