"""5월 7일 내부테스트 로그 분석 → Excel 리포트 (워크플로우/코드 위치 단위)."""

import re
import os
from urllib.parse import unquote_plus
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter

LOG_PATH = os.path.join(os.path.dirname(__file__), "..", "backend_logs.txt")
OUT_PATH = os.path.join(os.path.dirname(__file__), "..", "reports", "internal_test_20260507.xlsx")
os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)

# ── 1) 로그 읽기 ──────────────────────────────────────────────
with open(LOG_PATH, encoding="utf-8", errors="replace") as f:
    all_lines = f.readlines()

start_line = end_line = None
for i, line in enumerate(all_lines):
    if "2026-05-07" in line and start_line is None:
        start_line = i
    if "2026-05-08" in line and start_line is not None:
        end_line = i
        break
if end_line is None:
    end_line = len(all_lines)

may7_lines = all_lines[start_line:end_line]

# ── 2) 이벤트 파싱 ────────────────────────────────────────────
TS_RE = re.compile(r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})")
STREAM_RE = re.compile(r'"GET /api/chat/stream\?session_id=([^&"]+)&question=([^" ]+)')
TIMING_RE = re.compile(
    r"PIPELINE_TIMING total=(\d+)ms.*?follow_up=(\d+)ms rewrite=(\d+)ms"
    r" analyze=(\d+)ms search=(\d+)ms merge=(\d+)ms retry=(\d+)ms"
    r" generate=(\d+)ms validate=(\d+)ms.*?intent=(\w+) qt=(\w+) follow_up=(\S+)"
)
REWRITE_ORIG_RE = re.compile(r"follow-up\[(\w+)\] rewrite: '(.+?)'\s*→")
MISMATCH_RE = re.compile(r"answer-context mismatch(?: \(sync\))?: (.+)$")
LLM_ERR_RE = re.compile(r"LLM 생성 오류: (.+)$")
EMPTY_LLM_RE = re.compile(r"LLM 빈 응답: question='(.+?)'")

current_ts = None
events = []

for idx, raw in enumerate(may7_lines):
    line = raw.rstrip()
    m = TS_RE.match(line)
    if m:
        try:
            current_ts = datetime.strptime(m.group(1), "%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass

    m = STREAM_RE.search(line)
    if m:
        q_enc = m.group(2).split(" ")[0].rstrip('"')
        try:
            q = unquote_plus(q_enc).strip()
        except Exception:
            q = q_enc
        events.append({"idx": idx, "type": "query", "ts": current_ts,
                        "session": m.group(1), "query": q})
        continue

    m = TIMING_RE.search(line)
    if m:
        events.append({"idx": idx, "type": "timing", "ts": current_ts,
                        "total": int(m.group(1)),
                        "follow_up_ms": int(m.group(2)),
                        "rewrite_ms": int(m.group(3)),
                        "analyze_ms": int(m.group(4)),
                        "search_ms": int(m.group(5)),
                        "merge_ms": int(m.group(6)),
                        "retry_ms": int(m.group(7)),
                        "gen_ms": int(m.group(8)),
                        "val_ms": int(m.group(9)),
                        "intent": m.group(10),
                        "qt": m.group(11),
                        "follow_up": m.group(12)})
        continue

    m = REWRITE_ORIG_RE.search(line)
    if m:
        events.append({"idx": idx, "type": "followup_rewrite", "ts": current_ts,
                        "reason": m.group(1), "query": m.group(2).strip()})

    m = MISMATCH_RE.search(line)
    if m:
        events.append({"idx": idx, "type": "mismatch", "ts": current_ts,
                        "detail": m.group(1).strip()})

    m = LLM_ERR_RE.search(line)
    if m:
        events.append({"idx": idx, "type": "llm_error", "ts": current_ts,
                        "detail": m.group(1).strip()})

    m = EMPTY_LLM_RE.search(line)
    if m:
        events.append({"idx": idx, "type": "empty_answer", "ts": current_ts,
                        "query": m.group(1).strip()})

queries_ev = [e for e in events if e["type"] == "query"]
timings_ev = [e for e in events if e["type"] == "timing"]
followup_ev = [e for e in events if e["type"] == "followup_rewrite"]
mismatch_ev = [e for e in events if e["type"] == "mismatch"]
empty_ev = [e for e in events if e["type"] == "empty_answer"]

print(f"이벤트 수집: queries={len(queries_ev)}, timings={len(timings_ev)}, "
      f"followup={len(followup_ev)}, mismatch={len(mismatch_ev)}, empty={len(empty_ev)}")

# ── 3) 30개 유니크 질문 선택 ──────────────────────────────────
seen = set()
unique_queries = []
for q in queries_ev:
    key = q["query"].strip()
    if key and key not in seen:
        seen.add(key)
        unique_queries.append(q)

selected = unique_queries[:30]

# ── 4) query → timing 순차 1:1 매핑 ─────────────────────────
def match_seq(q_list, t_list):
    t_ptr = 0
    out = []
    for q in q_list:
        found = None
        while t_ptr < len(t_list):
            t = t_list[t_ptr]
            if t["idx"] > q["idx"]:
                found = t
                t_ptr += 1
                break
            t_ptr += 1
        out.append(found)
    return out

all_q_timings = match_seq(queries_ev, timings_ev)
q_idx_to_timing = {q["idx"]: t for q, t in zip(queries_ev, all_q_timings)}

# rewrite_map: 원본 쿼리 → follow_up reason
rewrite_map = {fw["query"].strip(): fw["reason"] for fw in followup_ev}

def near_mismatch(q_idx, window=80):
    for m in mismatch_ev:
        if q_idx < m["idx"] <= q_idx + window:
            return m
    return None

# ── 5) 워크플로우 단계 매핑 ───────────────────────────────────
# 각 문제 유형 → 책임 stage, 파일:라인, 함수, 로그 키
STAGE_INFO = {
    # 키 = 문제 유형
    "FOLLOWUP_FALSE_POSITIVE": {
        "stage": "Stage 0: follow-up 감지",
        "file_func": "app/pipeline/follow_up_detector.py:_has_subject_ko()",
        "log_key": "logger.info('follow-up[no_subject_short_ko] rewrite: ...')",
        "trigger": "주어 휴리스틱 정규식이 '조사+구두점' 종료 패턴 미인식 (예: '전공은?')",
        "fix": "commit 8ac6a65 — `(?=\\s|[?!.,]|$)` lookahead 추가 (배포 필요)",
    },
    "ANSWER_CONTEXT_MISMATCH": {
        "stage": "Phase 4 품질 게이트 (Stage 5.5)",
        "file_func": "app/pipeline/answer_units.py:verify_answer_against_context()",
        "log_key": "logger.warning('answer-context mismatch: ...')",
        "trigger": "LLM이 컨텍스트에 없는 숫자/URL 환각 생성 (credit:43/131, url:m.bufs.ac.kr 등)",
        "fix": "P0 패치 적용 후 재인덱싱 (그래프 메타 양방향 미러링 + 청크 ID SHA256)",
    },
    "OUT_OF_SCOPE": {
        "stage": "Stage 2: 검색 (search) + Stage 3: merge",
        "file_func": "app/pipeline/query_router.py:route_and_search() → context_merger:merge()",
        "log_key": "logger.info('라우팅: intent=GENERAL, vector=X, graph=Y') + context_confidence<0.3",
        "trigger": "ChromaDB·그래프에 매칭 데이터 없음 (식당·셔틀버스·주차·운동장 도메인 미포함)",
        "fix": "data/contacts/departments.json 또는 FAQ에 시설·교통 안내 추가",
    },
    "TIMEOUT_RISK": {
        "stage": "Stage 5: LLM 스트리밍 생성",
        "file_func": "app/pipeline/answer_generator.py:generate() (Ollama /api/chat)",
        "log_key": "logger.info('PIPELINE_TIMING ... generate=>29000ms')",
        "trigger": "LLM 응답 ≥29s — Ollama 모델 로드 콜드스타트 또는 컨텍스트 과대",
        "fix": "ctx_chars 축소 / ollama keep-alive 설정 / max_tokens 하향",
    },
    "REWRITE_DELAY": {
        "stage": "Stage 0.5: query_rewriter (LLM 호출)",
        "file_func": "app/pipeline/query_rewriter.py:rewrite() (gemma3:4b)",
        "log_key": "logger.info('PIPELINE_TIMING ... rewrite=>3000ms')",
        "trigger": "재작성 LLM 응답 ≥3s (rewrite_timeout_sec=0.8 초과)",
        "fix": "타임아웃 상향 또는 룰 기반 fallback 강화",
    },
    "RESPONSE_DELAY": {
        "stage": "Stage 5: 생성",
        "file_func": "app/pipeline/answer_generator.py:generate() + ChromaDB 검색 병목",
        "log_key": "PIPELINE_TIMING total > 15000ms",
        "trigger": "전체 응답시간 15s↑ (검색·생성 누적)",
        "fix": "search_ms 분석 / generate ctx_chars 검토",
    },
    "FORMAT_REQUEST": {
        "stage": "Stage 5: LLM 프롬프트",
        "file_func": "app/pipeline/answer_generator.py:SYSTEM_PROMPT",
        "log_key": "—",
        "trigger": "사용자가 표/그래프 형식 요청 — 현 프롬프트는 마크다운 표 지시 부재",
        "fix": "SYSTEM_PROMPT에 '여러 항목은 표로 정리' 룰 추가",
    },
    "PERSONAL_QUERY": {
        "stage": "Stage 1: query_analyzer (의도 미스매칭)",
        "file_func": "app/pipeline/query_analyzer.py:analyze()",
        "log_key": "intent=COURSE_INFO + 학번/분반 키워드",
        "trigger": "개인 수강시간표는 LMS·포털에서만 조회 가능 (RAG KB 범위 아님)",
        "fix": "PERSONAL_SCHEDULE intent 신설 → 포털 안내 답변으로 단락",
    },
    "OK": {
        "stage": "—",
        "file_func": "—",
        "log_key": "—",
        "trigger": "정상 응답",
        "fix": "—",
    },
}

SEVERITY = {
    "TIMEOUT_RISK": 5,
    "ANSWER_CONTEXT_MISMATCH": 4,
    "RESPONSE_DELAY": 3,
    "REWRITE_DELAY": 3,
    "OUT_OF_SCOPE": 2,
    "FOLLOWUP_FALSE_POSITIVE": 2,
    "FORMAT_REQUEST": 1,
    "PERSONAL_QUERY": 1,
    "OK": 0,
}

OUTSCOPE_KW = ["식당", "메뉴", "셔틀버스", "주차", "운동장", "와이파이"]
FORMAT_KW = ["표로 만들", "차트로", "그래프로"]
PERSONAL_KW = ["gyo0", "채플 2분반", "채플 02분반", "경찰학총론 수업"]

def classify(query, timing, mismatch, fu_reason):
    issues = []
    q = query.lower()

    if any(k in q for k in [k.lower() for k in OUTSCOPE_KW]):
        issues.append("OUT_OF_SCOPE")
    if any(k in q for k in [k.lower() for k in FORMAT_KW]):
        issues.append("FORMAT_REQUEST")
    if any(k in q for k in [k.lower() for k in PERSONAL_KW]):
        issues.append("PERSONAL_QUERY")

    if mismatch:
        issues.append("ANSWER_CONTEXT_MISMATCH")

    if timing:
        fu = timing.get("follow_up", fu_reason or "")
        total = timing.get("total", 0)
        rw_ms = timing.get("rewrite_ms", 0)

        if fu == "no_subject_short_ko" and "OUT_OF_SCOPE" not in issues:
            issues.append("FOLLOWUP_FALSE_POSITIVE")
        if rw_ms > 3000 and "FOLLOWUP_FALSE_POSITIVE" not in issues:
            issues.append("REWRITE_DELAY")
        if total > 29000:
            issues.append("TIMEOUT_RISK")
        elif total > 15000:
            issues.append("RESPONSE_DELAY")
    elif fu_reason == "no_subject_short_ko" and "OUT_OF_SCOPE" not in issues:
        issues.append("FOLLOWUP_FALSE_POSITIVE")

    if not issues:
        issues.append("OK")
    return issues

def sev(issues):
    return max(SEVERITY.get(i, 0) for i in issues)

# ── 6) 행 데이터 생성 ─────────────────────────────────────────
COLOR_RED = "FFD6D6"
COLOR_ORANGE = "FFEACC"
COLOR_YELLOW = "FFFACD"
COLOR_GREEN = "D6EED6"
COLOR_HEADER = "2F5496"
COLOR_WHITE = "FFFFFF"

def row_fill(s):
    if s >= 4: return COLOR_RED
    if s >= 3: return COLOR_ORANGE
    if s >= 2: return COLOR_YELLOW
    if s == 0: return COLOR_GREEN
    return COLOR_WHITE

rows = []
for i, q in enumerate(selected, 1):
    timing = q_idx_to_timing.get(q["idx"])
    mismatch = near_mismatch(q["idx"])
    fu_reason = rewrite_map.get(q["query"].strip(), "")
    issues = classify(q["query"], timing, mismatch, fu_reason)

    # 가장 심각한 이슈를 "주된 실패 stage"로 선택
    primary = max(issues, key=lambda x: SEVERITY.get(x, 0))
    s = sev(issues)

    kst_str = "-"
    if q["ts"]:
        kst = q["ts"] + timedelta(hours=9)
        kst_str = kst.strftime("%H:%M:%S")

    stage_info = STAGE_INFO[primary]
    rows.append({
        "no": i,
        "query": q["query"].replace("\n", "").strip(),
        "kst": kst_str,
        "intent": timing["intent"] if timing else "-",
        "qt": timing["qt"] if timing else "-",
        "follow_up": (timing["follow_up"] if timing else fu_reason) or "-",
        "total_ms": timing["total"] if timing else "-",
        "rewrite_ms": timing["rewrite_ms"] if timing else "-",
        "search_ms": timing["search_ms"] if timing else "-",
        "merge_ms": timing["merge_ms"] if timing else "-",
        "gen_ms": timing["gen_ms"] if timing else "-",
        "val_ms": timing["val_ms"] if timing else "-",
        "warning": mismatch["detail"] if mismatch else "",
        "issues": issues,
        "primary": primary,
        "stage": stage_info["stage"],
        "file_func": stage_info["file_func"],
        "log_key": stage_info["log_key"],
        "trigger": stage_info["trigger"],
        "fix": stage_info["fix"],
        "sev": s,
        "fill": row_fill(s),
    })

print(f"\n[분석된 30개 질문 — 주요 실패 stage]")
stages = Counter(r["primary"] for r in rows)
for st, cnt in stages.most_common():
    print(f"  {st:30s} {cnt:2d}건  ({STAGE_INFO[st]['file_func']})")

# ── 7) Excel 생성 ─────────────────────────────────────────────
wb = openpyxl.Workbook()

def fnt(bold=False, sz=10, color="000000"):
    return Font(name="맑은 고딕", bold=bold, size=sz, color=color)

def fill(c):
    return PatternFill("solid", fgColor=c)

thin = Side(style="thin", color="BFBFBF")
BRD = Border(left=thin, right=thin, top=thin, bottom=thin)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── 시트 1: 질문별 워크플로우 분석 ────────────────────────────
ws = wb.active
ws.title = "30개 질문 · 워크플로우 분석"

COLS = [
    ("No.", 5, AC),
    ("질문 내용", 38, AL),
    ("KST 시각", 10, AC),
    ("의도", 14, AC),
    ("질문유형", 11, AC),
    ("Follow-up 판정", 22, AC),
    ("총 ms", 9, AC),
    ("재작성 ms", 10, AC),
    ("검색 ms", 9, AC),
    ("merge ms", 9, AC),
    ("생성 ms", 9, AC),
    ("검증 ms", 9, AC),
    ("주요 실패 stage", 28, AL),
    ("실패 위치 (파일:함수)", 50, AL),
    ("로그 키", 50, AL),
    ("트리거 원인", 60, AL),
    ("권장 조치", 55, AL),
]
for c, (h, w, _) in enumerate(COLS, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = fnt(bold=True, color="FFFFFF")
    cell.fill = fill(COLOR_HEADER)
    cell.alignment = AC
    cell.border = BRD
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
ws.row_dimensions[1].height = 32

NUM_COLS = {1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12}
for r, row in enumerate(rows, 2):
    vals = [
        row["no"], row["query"], row["kst"], row["intent"], row["qt"],
        row["follow_up"], row["total_ms"], row["rewrite_ms"], row["search_ms"],
        row["merge_ms"], row["gen_ms"], row["val_ms"],
        row["stage"], row["file_func"], row["log_key"], row["trigger"], row["fix"],
    ]
    f = fill(row["fill"])
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = fnt(sz=9)
        cell.fill = f
        cell.alignment = AC if c in NUM_COLS else AL
        cell.border = BRD
    ws.row_dimensions[r].height = 48

ws.freeze_panes = "C2"

# ── 시트 2: 실패 stage별 집계 ─────────────────────────────────
ws2 = wb.create_sheet("실패 stage별 집계")

primary_counts = Counter(r["primary"] for r in rows)

COLS2 = [
    ("실패 코드", 26),
    ("건수", 7),
    ("Pipeline Stage", 30),
    ("코드 위치 (파일:함수)", 50),
    ("핵심 로그 키", 55),
    ("트리거 원인", 60),
    ("권장 조치", 55),
    ("예시 질문", 35),
]

for c, (h, w) in enumerate(COLS2, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = fnt(bold=True, color="FFFFFF")
    cell.fill = fill(COLOR_HEADER)
    cell.alignment = AC
    cell.border = BRD
    ws2.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
ws2.row_dimensions[1].height = 32

r2 = 2
for code, cnt in sorted(primary_counts.items(), key=lambda x: -SEVERITY.get(x[0], 0)):
    info = STAGE_INFO[code]
    examples = [row["query"] for row in rows if row["primary"] == code][:2]
    s = SEVERITY.get(code, 0)
    f = fill(row_fill(s))

    vals = [
        code, cnt, info["stage"], info["file_func"], info["log_key"],
        info["trigger"], info["fix"], " / ".join(examples),
    ]
    for c, val in enumerate(vals, 1):
        cell = ws2.cell(row=r2, column=c, value=val)
        cell.font = fnt(sz=9)
        cell.fill = f
        cell.alignment = AC if c == 2 else AL
        cell.border = BRD
    ws2.row_dimensions[r2].height = 60
    r2 += 1

# 색상 범례
r2 += 1
ws2.cell(row=r2, column=1, value="[색상 범례]").font = fnt(bold=True)
r2 += 1
for color_c, desc in [
    (COLOR_RED, "빨강 ── 심각도 4-5: 타임아웃 / 답변-컨텍스트 불일치"),
    (COLOR_ORANGE, "주황 ── 심각도 3: 응답·재작성 지연"),
    (COLOR_YELLOW, "노랑 ── 심각도 2: follow_up 오검출 / 범위외"),
    (COLOR_GREEN, "초록 ── 정상 (심각도 0)"),
]:
    cell = ws2.cell(row=r2, column=1, value=f"  {desc}")
    cell.fill = fill(color_c)
    cell.font = fnt()
    ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=8)
    ws2.row_dimensions[r2].height = 18
    r2 += 1

# ── 시트 3: Pipeline Stage 다이어그램 ──────────────────────────
ws3 = wb.create_sheet("Pipeline Stage 매핑")
ws3.column_dimensions["A"].width = 6
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 50
ws3.column_dimensions["D"].width = 40
ws3.column_dimensions["E"].width = 18

PIPELINE_STAGES = [
    ("0", "follow_up 감지", "app/pipeline/follow_up_detector.py:detect()",
     "현재 쿼리가 직전 턴 종속인지 휴리스틱 판정", "follow_up_detector"),
    ("0.5", "query 재작성 (LLM)", "app/pipeline/query_rewriter.py:rewrite()",
     "follow-up이면 gemma3:4b로 self-contained 쿼리 생성", "query_rewriter"),
    ("1", "질문 분석", "app/pipeline/query_analyzer.py:analyze()",
     "intent·question_type·entities·lang 추출", "query_analyzer"),
    ("2", "라우팅·검색", "app/pipeline/query_router.py:route_and_search()",
     "Vector(ChromaDB) + Graph(NetworkX) 동시 검색 + reranker", "query_router"),
    ("3", "컨텍스트 병합", "app/pipeline/context_merger.py:merge()",
     "RRF 융합 + evidence_slicing + context_confidence 계산", "context_merger"),
    ("4", "P4 재시도 (조건부)", "app/pipeline/answer_generator.py:rewrite_query()",
     "confidence<0.3 + direct_answer 없을 때 1회 재시도", "P4 retry"),
    ("5", "LLM 생성 (스트리밍)", "app/pipeline/answer_generator.py:generate()",
     "Ollama /api/chat (gemma4:26b) 스트리밍", "[generator-IN/OUT]"),
    ("5.5", "Phase 4 품질 게이트 (KO)", "app/pipeline/answer_units.py:verify_answer_against_context()",
     "답변에 컨텍스트 외 숫자/URL 환각 탐지", "answer-context mismatch"),
    ("6", "응답 검증", "app/pipeline/response_validator.py:validate()",
     "출처 표기·환각 패턴 최종 검사", "[validator-IN/OUT]"),
]

for c, h in enumerate(["#", "Stage 명", "파일:함수", "역할", "로그 키 prefix"], 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.font = fnt(bold=True, color="FFFFFF")
    cell.fill = fill(COLOR_HEADER)
    cell.alignment = AC
    cell.border = BRD
ws3.row_dimensions[1].height = 28

for r3, (num, name, ff, role, logk) in enumerate(PIPELINE_STAGES, 2):
    vals = [num, name, ff, role, logk]
    for c, val in enumerate(vals, 1):
        cell = ws3.cell(row=r3, column=c, value=val)
        cell.font = fnt(sz=9)
        cell.alignment = AC if c in (1, 5) else AL
        cell.border = BRD
    ws3.row_dimensions[r3].height = 30

# ── 시트 4: 신규 로그 사양 ────────────────────────────────────
ws4 = wb.create_sheet("신규 로그 사양")
ws4.column_dimensions["A"].width = 32
ws4.column_dimensions["B"].width = 95

NEW_LOGS = [
    ("[generator-IN]",
     "sid=<8자> model=<llm> prompt_ctx_chars=<n> intent=<...> qt=<...> lang=<ko/en> history_turns=<n> search_n=<n>\n"
     "→ 위치: backend/routers/chat.py Stage 5 진입 직전"),
    ("[generator-IN-cand]",
     "sid=<8자> #<i>[<source_filename>:p<page>|s=<score>] <text 50자>\n"
     "→ 검색 후보 8개(최대 16개) 각각 1줄. source/page/score/텍스트 미리보기 포함."),
    ("[generator-OUT]",
     "sid=<8자> answer_chars=<n> elapsed_ms=<ms> preview='<답변 200자>'\n"
     "→ LLM 생성 완료 직후. answer_chars=17 같은 짧은 응답의 실제 텍스트를 preview로 확인 가능."),
    ("[validator-IN]",
     "sid=<8자> answer_chars=<n> ctx_chars=<n> search_n=<n>\n"
     "→ response_validator.validate() 호출 직전"),
    ("[validator-OUT]",
     "sid=<8자> passed=<bool> warnings=<n> elapsed_ms=<ms>\n"
     "→ 검증 완료 직후"),
    ("answer-context mismatch",
     "기존 로그 유지 — Phase 4 환각 탐지 경고"),
    ("PIPELINE_TIMING",
     "기존 로그 유지 — total + 각 stage별 ms + intent/qt/follow_up"),
]

for c, h in enumerate(["로그 키", "포맷 / 위치"], 1):
    cell = ws4.cell(row=1, column=c, value=h)
    cell.font = fnt(bold=True, color="FFFFFF")
    cell.fill = fill(COLOR_HEADER)
    cell.alignment = AC
    cell.border = BRD
ws4.row_dimensions[1].height = 26

for r4, (k, v) in enumerate(NEW_LOGS, 2):
    ws4.cell(row=r4, column=1, value=k).font = fnt(sz=10, bold=True)
    ws4.cell(row=r4, column=2, value=v).font = fnt(sz=9)
    ws4.cell(row=r4, column=1).alignment = AL
    ws4.cell(row=r4, column=2).alignment = AL
    ws4.cell(row=r4, column=1).border = BRD
    ws4.cell(row=r4, column=2).border = BRD
    ws4.row_dimensions[r4].height = 50

# 도커 명령 안내
r4 = len(NEW_LOGS) + 4
ws4.cell(row=r4, column=1, value="[도커 로그 조회 명령]").font = fnt(bold=True, sz=11)
r4 += 1
DOCKER_CMDS = [
    "docker compose -f docker/docker-compose.yml logs -f backend",
    "docker compose -f docker/docker-compose.yml logs backend --tail=200 | grep '[generator-IN'",
    "docker compose -f docker/docker-compose.yml logs backend | grep -E '\\[generator-(IN|OUT)\\]|\\[validator-(IN|OUT)\\]|answer-context mismatch'",
    "# 단일 세션 추적 (sid 8자):",
    "docker compose -f docker/docker-compose.yml logs backend | grep 'sid=9f2db6c9'",
]
for cmd in DOCKER_CMDS:
    cell = ws4.cell(row=r4, column=1, value=cmd)
    cell.font = Font(name="Consolas", size=9)
    cell.alignment = AL
    ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=2)
    ws4.row_dimensions[r4].height = 20
    r4 += 1

wb.save(OUT_PATH)
print(f"\n✓ Excel 저장: {OUT_PATH}")
print(f"  시트1 '30개 질문 · 워크플로우 분석': {len(rows)}행 × 17열")
print(f"  시트2 '실패 stage별 집계': {len(primary_counts)}개 stage")
print(f"  시트3 'Pipeline Stage 매핑': {len(PIPELINE_STAGES)}개 stage")
print(f"  시트4 '신규 로그 사양': {len(NEW_LOGS)}개 로그 키")
