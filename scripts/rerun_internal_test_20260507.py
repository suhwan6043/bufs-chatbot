"""5/7 내부테스트 30개 질문 재실행 (재인제스트 후) + 이전 결과 비교."""

import json
import os
import time
import sys
import uuid
from datetime import datetime
from pathlib import Path

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
PREV_REPORT = ROOT / "reports" / "internal_test_20260507.xlsx"
OUT_REPORT = ROOT / "reports" / "internal_test_20260507_rerun.xlsx"
OUT_RAW = ROOT / "reports" / "internal_test_20260507_rerun.jsonl"
BASE_URL = os.getenv("BACKEND_BASE_URL", "http://localhost:8000")

OUT_REPORT.parent.mkdir(parents=True, exist_ok=True)

# ── 1) 이전 결과 로드 ─────────────────────────────────────────
print(f"[1/4] 이전 결과 로드: {PREV_REPORT}")
wb_prev = openpyxl.load_workbook(PREV_REPORT)
ws_prev = wb_prev["30개 질문 · 워크플로우 분석"]

prev_rows = []
for r in range(2, ws_prev.max_row + 1):
    prev_rows.append({
        "no": ws_prev.cell(r, 1).value,
        "query": ws_prev.cell(r, 2).value,
        "kst_prev": ws_prev.cell(r, 3).value,
        "intent_prev": ws_prev.cell(r, 4).value,
        "qt_prev": ws_prev.cell(r, 5).value,
        "follow_up_prev": ws_prev.cell(r, 6).value,
        "total_ms_prev": ws_prev.cell(r, 7).value,
        "stage_prev": ws_prev.cell(r, 13).value,
        "fix_prev": ws_prev.cell(r, 17).value,
    })
print(f"   이전 30개 질문 로드 완료")

# ── 2) 백엔드 헬스체크 ────────────────────────────────────────
print(f"[2/4] 백엔드 헬스체크: {BASE_URL}")
try:
    h = httpx.get(f"{BASE_URL}/api/health", timeout=5)
    h.raise_for_status()
    print(f"   ✓ backend OK: {h.json()}")
except Exception as e:
    print(f"   ✗ backend 응답 없음: {e}")
    sys.exit(1)

# ── 3) 30개 질문 재실행 ───────────────────────────────────────
print(f"[3/4] 30개 질문 재실행 (각 질문별 새 세션, X-Test-Mode=1)")

# 새 컬렉션·새 그래프가 적용됐는지 간단 진단 (질문 1개 빠른 호출)
results = []
client = httpx.Client(base_url=BASE_URL, timeout=180,
                      headers={"X-Test-Mode": "1"})

for i, row in enumerate(prev_rows, 1):
    sid = uuid.uuid4().hex[:12]
    t0 = time.monotonic()
    print(f"   [{i:2d}/30] {row['query'][:50]}", end=" ", flush=True)
    try:
        r = client.post("/api/chat",
                        params={"session_id": sid, "question": row["query"]})
        r.raise_for_status()
        d = r.json()
        elapsed_ms = int((time.monotonic() - t0) * 1000)
        answer = d.get("answer", "")
        intent = d.get("intent", "")
        srcs = d.get("source_urls", []) or []
        rresults = d.get("results", []) or []
        timing = d.get("timing", {}) or {}
        rec = {
            "no": row["no"],
            "query": row["query"],
            "sid": sid,
            "answer": answer,
            "answer_chars": len(answer),
            "intent": intent,
            "source_urls_n": len(srcs),
            "results_n": len(rresults),
            "elapsed_ms": elapsed_ms,
            "timing": timing,
            "results_summary": [
                {"page": x.get("page_number"), "source": x.get("source"),
                 "score": x.get("score"), "doc_type": (x.get("metadata") or {}).get("doc_type")}
                for x in rresults[:8]
            ],
            "ts_utc": datetime.utcnow().isoformat() + "Z",
            "error": None,
        }
        print(f"  → {len(answer):4d}자  intent={intent:18s}  {elapsed_ms/1000:.1f}s")
    except Exception as e:
        elapsed_ms = int((time.monotonic() - t0) * 1000)
        rec = {
            "no": row["no"], "query": row["query"], "sid": sid,
            "answer": "", "answer_chars": 0, "intent": "",
            "source_urls_n": 0, "results_n": 0, "elapsed_ms": elapsed_ms,
            "timing": {}, "results_summary": [],
            "ts_utc": datetime.utcnow().isoformat() + "Z",
            "error": f"{type(e).__name__}: {e}",
        }
        print(f"  ✗ ERROR: {rec['error']}")
    results.append(rec)

client.close()

with open(OUT_RAW, "w", encoding="utf-8") as f:
    for rec in results:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")
print(f"   원본 응답 저장: {OUT_RAW}")

# ── 4) 비교 분석 + Excel ──────────────────────────────────────
print(f"[4/4] 비교 분석 + Excel 생성")

# 문제 유형 판단 (개략, intent/answer 기반)
OUTSCOPE_KW = ["식당", "메뉴", "셔틀버스", "주차", "운동장", "와이파이"]
REFUSAL_PHRASES = [
    "찾을 수 없", "정확히 확인하지 못", "문의하시기 바랍니다",
    "couldn't find", "no relevant",
]


def classify_after(rec, prev_stage):
    issues = []
    a = rec.get("answer", "") or ""
    q = rec.get("query", "") or ""
    total = rec.get("elapsed_ms", 0)
    err = rec.get("error")
    if err:
        issues.append("ERROR")
        return issues
    is_refusal = any(p in a for p in REFUSAL_PHRASES)
    out_scope = any(k in q for k in OUTSCOPE_KW)
    if out_scope and is_refusal:
        issues.append("OUT_OF_SCOPE(refused)")  # 정상적인 거절
    elif out_scope:
        issues.append("OUT_OF_SCOPE(answered)")
    elif is_refusal:
        issues.append("REFUSED")
    if total >= 29000:
        issues.append("TIMEOUT_RISK")
    elif total >= 15000:
        issues.append("RESPONSE_DELAY_15s")
    if rec.get("answer_chars", 0) < 30 and not is_refusal:
        issues.append("VERY_SHORT_ANSWER")
    if not issues:
        issues.append("OK")
    return issues


def compare(prev_stage, after_issues):
    if "ERROR" in after_issues:
        return "ERROR(새 오류)"
    if prev_stage and "FOLLOWUP_FALSE_POSITIVE" in prev_stage and "OK" in after_issues:
        return "개선 (follow_up 패치 + 재인제스트 효과)"
    if prev_stage and "ANSWER_CONTEXT_MISMATCH" in prev_stage and "OK" in after_issues:
        return "개선 (재인제스트로 컨텍스트 불일치 해소)"
    if prev_stage and "TIMEOUT_RISK" in prev_stage and "OK" in after_issues:
        return "개선 (응답시간 정상화)"
    if "OK" in after_issues and prev_stage == "—":
        return "변동 없음(여전히 정상)"
    if prev_stage and "OK" in after_issues:
        return "개선"
    if "TIMEOUT_RISK" in after_issues:
        return "악화 (새로 타임아웃 위험)"
    if "RESPONSE_DELAY_15s" in after_issues:
        return "악화 (응답 지연)"
    if "REFUSED" in after_issues and prev_stage and "OK" in prev_stage:
        return "악화 (이전엔 정상, 지금은 거절)"
    return "동일/판단보류"


# 점수 계산 — 직관적인 정상률
n_ok_before = sum(1 for r in prev_rows if r["stage_prev"] == "—")
n_ok_after = sum(1 for r in results if "OK" in classify_after(r, ""))

# Excel 작성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "재실행 결과 비교"

COLOR_HEADER = "2F5496"
COLOR_BETTER = "D6EED6"
COLOR_WORSE = "FFD6D6"
COLOR_SAME = "FFFFFF"
COLOR_NEUTRAL = "FFFACD"
thin = Side(style="thin", color="BFBFBF")
BRD = Border(left=thin, right=thin, top=thin, bottom=thin)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left", vertical="center", wrap_text=True)


def fnt(bold=False, sz=10, color="000000"):
    return Font(name="맑은 고딕", bold=bold, size=sz, color=color)


def fill(c):
    return PatternFill("solid", fgColor=c)


HEADERS = [
    ("No.", 5),
    ("질문", 36),
    ("이전 stage", 25),
    ("재실행 intent", 18),
    ("재실행 응답시간(ms)", 14),
    ("재실행 응답 길이", 12),
    ("재실행 source_urls", 12),
    ("재실행 results", 10),
    ("재실행 문제 분류", 28),
    ("변화", 28),
    ("재실행 답변 미리보기(120자)", 50),
]
for c, (h, w) in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = fnt(bold=True, color="FFFFFF")
    cell.fill = fill(COLOR_HEADER)
    cell.alignment = AC
    cell.border = BRD
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
ws.row_dimensions[1].height = 32

NUM_COLS = {1, 5, 6, 7, 8}

for i, (prev, rec) in enumerate(zip(prev_rows, results), 2):
    issues = classify_after(rec, prev["stage_prev"])
    change = compare(prev["stage_prev"], issues)
    if "개선" in change:
        f = fill(COLOR_BETTER)
    elif "악화" in change or "ERROR" in change:
        f = fill(COLOR_WORSE)
    elif "동일" in change or "변동 없음" in change:
        f = fill(COLOR_SAME)
    else:
        f = fill(COLOR_NEUTRAL)

    preview = (rec.get("answer", "") or "").replace("\n", " | ")[:120]
    vals = [
        prev["no"], prev["query"],
        prev["stage_prev"] or "—",
        rec.get("intent", "") or "?",
        rec.get("elapsed_ms", "-"),
        rec.get("answer_chars", "-"),
        rec.get("source_urls_n", "-"),
        rec.get("results_n", "-"),
        " / ".join(issues),
        change,
        preview,
    ]
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=i, column=c, value=val)
        cell.font = fnt(sz=9)
        cell.fill = f
        cell.alignment = AC if c in NUM_COLS else AL
        cell.border = BRD
    ws.row_dimensions[i].height = 38
ws.freeze_panes = "C2"

# ── 시트 2: 요약 ─────────────────────────────────────────────
ws2 = wb.create_sheet("재실행 요약")
ws2.column_dimensions["A"].width = 36
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 38

from collections import Counter
prev_stage_cnt = Counter(r["stage_prev"] or "—" for r in prev_rows)
after_issues_flat = []
for rec in results:
    after_issues_flat.extend(classify_after(rec, ""))
after_cnt = Counter(after_issues_flat)
change_cnt = Counter(
    compare(prev["stage_prev"], classify_after(rec, prev["stage_prev"]))
    for prev, rec in zip(prev_rows, results)
)
total_ms_after = [r["elapsed_ms"] for r in results if not r.get("error")]
avg_ms_after = (sum(total_ms_after) // len(total_ms_after)) if total_ms_after else 0
prev_ms = [r["total_ms_prev"] for r in prev_rows if isinstance(r["total_ms_prev"], int)]
avg_ms_prev = (sum(prev_ms) // len(prev_ms)) if prev_ms else 0

stats = [
    ("=== 정상률 ===", "", ""),
    ("이전 정상 (stage='—')", str(n_ok_before) + " / 30",
     f"{n_ok_before/30*100:.1f}%"),
    ("재실행 정상 (OK)", str(n_ok_after) + " / 30",
     f"{n_ok_after/30*100:.1f}%"),
    ("개선폭", f"{n_ok_after - n_ok_before:+d}건",
     f"{(n_ok_after - n_ok_before)/30*100:+.1f}pp"),
    ("", "", ""),
    ("=== 응답 시간 ===", "", ""),
    ("이전 평균 응답시간 (ms)", str(avg_ms_prev), ""),
    ("재실행 평균 응답시간 (ms)", str(avg_ms_after), ""),
    ("응답시간 변화", f"{avg_ms_after - avg_ms_prev:+d} ms",
     f"{(avg_ms_after - avg_ms_prev) / avg_ms_prev * 100:+.1f}%" if avg_ms_prev else "-"),
    ("", "", ""),
    ("=== 이전 stage 분포 ===", "", ""),
]
for s, c in sorted(prev_stage_cnt.items(), key=lambda x: -x[1]):
    stats.append((s, str(c), ""))
stats.append(("", "", ""))
stats.append(("=== 재실행 문제 분포 ===", "", ""))
for s, c in sorted(after_cnt.items(), key=lambda x: -x[1]):
    stats.append((s, str(c), ""))
stats.append(("", "", ""))
stats.append(("=== 변화 유형 ===", "", ""))
for s, c in sorted(change_cnt.items(), key=lambda x: -x[1]):
    stats.append((s, str(c), ""))

for r2, (k, v, pct) in enumerate(stats, 1):
    is_header = "===" in (k or "")
    ws2.cell(row=r2, column=1, value=k).font = fnt(bold=is_header)
    ws2.cell(row=r2, column=2, value=v).font = fnt(bold=is_header)
    ws2.cell(row=r2, column=3, value=pct).font = fnt(bold=is_header)
    if is_header:
        ws2.cell(row=r2, column=1).fill = fill(COLOR_HEADER)
        ws2.cell(row=r2, column=1).font = fnt(bold=True, color="FFFFFF")

wb.save(OUT_REPORT)
print(f"\n✓ 비교 리포트 저장: {OUT_REPORT}")
print(f"\n[정상률]   {n_ok_before}/30 ({n_ok_before/30*100:.1f}%) → "
      f"{n_ok_after}/30 ({n_ok_after/30*100:.1f}%)  "
      f"({(n_ok_after-n_ok_before)/30*100:+.1f}pp)")
print(f"[응답시간] 평균 {avg_ms_prev}ms → {avg_ms_after}ms")
print("\n[변화 유형 분포]")
for k, v in sorted(change_cnt.items(), key=lambda x: -x[1]):
    print(f"  {k}: {v}건")
